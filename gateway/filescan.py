"""File PII scanning: extract text from an uploaded file (with OCR for
images and scanned PDFs), run it through core/detect + core/policy, and
return the redacted text plus a findings report.

Text-based formats (txt/pdf/docx/xlsx/csv) extract directly. Image formats
(png/jpg/jpeg/tiff/bmp/webp) and image-only ("scanned") PDF pages fall back
to Tesseract OCR via pytesseract -- so a photo of a document or a scanned
contract still gets its PII detected and masked before it ever reaches a
model.

No vault involvement (mirrors filescan-worker/scan.py): a file drop is a
standalone governance surface, not a chat session with a rehydration story.
REVERSIBLE and BLOCK spans are both masked in the returned text.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from typing import Any

from contracts.spans import TextUnit, TextUnitLocator
from detect.pipeline import DetectionPipeline
from policy.engine import evaluate
from policy.schema import Policy

_TEXT_CONTENT_TYPES = {
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "text/csv": "csv",
    "text/plain": "txt",
}
_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "tif", "tiff", "bmp", "webp", "gif"}
_TEXT_EXTENSIONS = {"pdf", "docx", "xlsx", "csv", "txt", "md", "log", "json"}

# Below this many extracted characters a PDF page is treated as image-only
# (scanned) and sent to OCR instead of trusting the near-empty text layer.
_SCANNED_PAGE_TEXT_THRESHOLD = 12


class UnsupportedFileTypeError(Exception):
    def __init__(self, content_type: str | None, filename: str | None) -> None:
        super().__init__(f"unsupported file type: content_type={content_type!r} filename={filename!r}")


class OcrUnavailableError(Exception):
    """Raised when an image/scanned page needs OCR but the Tesseract binary
    isn't installed (pip installs pytesseract, but the engine itself is a
    system package: `sudo apt install tesseract-ocr`)."""


def detect_kind(filename: str | None, content_type: str | None) -> str:
    if content_type in _TEXT_CONTENT_TYPES:
        return _TEXT_CONTENT_TYPES[content_type]
    if content_type and content_type.startswith("image/"):
        return "image"
    # Any text/* subtype (text/plain, text/markdown, text/x-log, ...) --
    # covers extensionless text files the OS reports as text/*.
    if content_type and content_type.startswith("text/"):
        return "txt"
    ext = filename.rsplit(".", 1)[-1].lower() if filename and "." in filename else ""
    if ext in _IMAGE_EXTENSIONS:
        return "image"
    if ext in _TEXT_EXTENSIONS:
        return "txt" if ext in ("md", "log", "json") else ext
    raise UnsupportedFileTypeError(content_type, filename)


def _ocr_image_bytes(png_bytes: bytes) -> str:
    try:
        import pytesseract
        from PIL import Image
    except ImportError as err:  # pragma: no cover
        raise OcrUnavailableError("pytesseract/Pillow not installed") from err
    try:
        img = Image.open(io.BytesIO(png_bytes))
        return pytesseract.image_to_string(img)
    except pytesseract.TesseractNotFoundError as err:
        raise OcrUnavailableError(
            "Tesseract OCR engine not found. Install it with: sudo apt install tesseract-ocr"
        ) from err


def _extract_pdf(data: bytes) -> list[tuple[str, str]]:
    import fitz  # PyMuPDF -- text extraction + page rendering, no external binary

    units: list[tuple[str, str]] = []
    with fitz.open(stream=data, filetype="pdf") as doc:
        for i, page in enumerate(doc):
            text = page.get_text().strip()
            if len(text) < _SCANNED_PAGE_TEXT_THRESHOLD:
                # Image-only / scanned page -> render at 2x and OCR it.
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                text = _ocr_image_bytes(pix.tobytes("png")).strip()
            if text:
                units.append((f"page {i + 1}", text))
    return units


def _extract_docx(data: bytes) -> list[tuple[str, str]]:
    from docx import Document as DocxDocument

    doc = DocxDocument(io.BytesIO(data))
    return [(f"paragraph {i + 1}", p.text) for i, p in enumerate(doc.paragraphs) if p.text.strip()]


def _extract_xlsx(data: bytes) -> list[tuple[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    units: list[tuple[str, str]] = []
    for ws in wb.worksheets:
        for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [str(c) for c in row if c is not None]
            if cells:
                units.append((f"{ws.title}!row {r}", " ".join(cells)))
    return units


def _extract_csv(data: bytes) -> list[tuple[str, str]]:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [(f"row {r}", " ".join(row)) for r, row in enumerate(reader, start=1) if any(row)]


def extract_units(data: bytes, kind: str) -> list[tuple[str, str]]:
    """Returns [(location, text), ...]. `location` is a human-readable
    pointer (page/paragraph/row) for the report -- never a sensitive value."""
    if kind == "pdf":
        return _extract_pdf(data)
    if kind == "docx":
        return _extract_docx(data)
    if kind == "xlsx":
        return _extract_xlsx(data)
    if kind == "csv":
        return _extract_csv(data)
    if kind == "txt":
        text = data.decode("utf-8", errors="replace")
        return [("document", text)] if text.strip() else []
    if kind == "image":
        text = _ocr_image_bytes(data).strip()
        return [("image", text)] if text else []
    raise UnsupportedFileTypeError(None, kind)


@dataclass
class FileFinding:
    location: str
    label: str
    action: str


@dataclass
class FileScanResult:
    filename: str
    kind: str
    policy_id: str
    units_scanned: int
    span_counts_by_label: dict[str, int]
    blocked: bool
    redacted_text: str
    findings: list[FileFinding] = field(default_factory=list)


def scan_file(data: bytes, filename: str, kind: str, pipeline: DetectionPipeline, policy: Policy) -> FileScanResult:
    raw_units = extract_units(data, kind)

    text_units: list[TextUnit] = []
    location_by_unit_id: dict[str, str] = {}
    text_by_unit_id: dict[str, str] = {}
    for i, (location, text) in enumerate(raw_units):
        unit_id = f"u{i}"
        text_units.append(TextUnit(
            unit_id=unit_id, role="user", text=text,
            locator=TextUnitLocator(surface="file_field", path=f"{filename}#{location}"),
            turn_index=i, direction="input",
        ))
        location_by_unit_id[unit_id] = location
        text_by_unit_id[unit_id] = text

    spans = pipeline.run(text_units, policy_ctx=policy) if text_units else []
    decisions = evaluate(spans, policy)

    span_counts: dict[str, int] = {}
    findings: list[FileFinding] = []
    blocked = False
    decisions_by_unit: dict[str, list[Any]] = {}
    for d in decisions:
        decisions_by_unit.setdefault(d.span.unit_id, []).append(d)
        span_counts[d.span.label.value] = span_counts.get(d.span.label.value, 0) + 1
        if d.action.value == "BLOCK":
            blocked = True
        findings.append(FileFinding(
            location=location_by_unit_id[d.span.unit_id], label=d.span.label.value, action=d.action.value,
        ))

    masked_parts: list[str] = []
    for unit in text_units:
        text = text_by_unit_id[unit.unit_id]
        unit_decisions = sorted(decisions_by_unit.get(unit.unit_id, []), key=lambda d: d.span.start)
        out: list[str] = []
        last_end = 0
        for d in unit_decisions:
            span = d.span
            if span.start < last_end:
                continue
            out.append(text[last_end:span.start])
            if d.action.value in ("BLOCK", "REVERSIBLE"):
                out.append(f"[REDACTED_{span.label.value}]")
            else:
                out.append(span.text)
            last_end = span.end
        out.append(text[last_end:])
        masked_parts.append("".join(out))

    return FileScanResult(
        filename=filename, kind=kind, policy_id=policy.policy_id, units_scanned=len(text_units),
        span_counts_by_label=span_counts, blocked=blocked,
        redacted_text="\n\n".join(masked_parts), findings=findings,
    )
